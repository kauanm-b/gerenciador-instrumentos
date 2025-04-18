"""
Módulo responsável pela geração das planilhas Excel.
"""
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelManager:
    """Classe responsável pela geração das planilhas Excel."""
    
    def __init__(self, pasta_saida: Path):
        """
        Inicializa o gerenciador de planilhas Excel.
        
        Args:
            pasta_saida: Pasta onde as planilhas serão salvas.
        """
        self.pasta_saida = pasta_saida
        self.pasta_saida.mkdir(parents=True, exist_ok=True)
        
        # Cores
        self.cor_azul = "1F4E78"
        self.cor_verde = "00B050"
        self.cor_vermelho = "FF0000"
        self.cor_amarelo = "FFC000"
        
        # Fontes
        self.fonte_titulo = Font(name="Arial", size=14, bold=True)
        self.fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.fonte_dados = Font(name="Arial", size=10)
        
        # Preenchimentos
        self.preenchimento_cabecalho = PatternFill(
            start_color=self.cor_azul,
            end_color=self.cor_azul,
            fill_type="solid"
        )
        self.preenchimento_verde = PatternFill(
            start_color=self.cor_verde,
            end_color=self.cor_verde,
            fill_type="solid"
        )
        self.preenchimento_vermelho = PatternFill(
            start_color=self.cor_vermelho,
            end_color=self.cor_vermelho,
            fill_type="solid"
        )
        self.preenchimento_amarelo = PatternFill(
            start_color=self.cor_amarelo,
            end_color=self.cor_amarelo,
            fill_type="solid"
        )
        
        # Alinhamentos
        self.alinhamento_central = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        
        # Bordas
        self.borda_fina = Border(
            left=Border(style="thin"),
            right=Border(style="thin"),
            top=Border(style="thin"),
            bottom=Border(style="thin")
        )
    
    def criar_planilha_geral(self, dados: pd.DataFrame) -> Path:
        """
        Cria a planilha geral com todos os instrumentos.
        
        Args:
            dados: DataFrame com os dados dos instrumentos.
            
        Returns:
            Caminho do arquivo Excel gerado.
        """
        # Cria o arquivo Excel
        nome_arquivo = "Relação de Instrumentos.xlsx"
        caminho_arquivo = self.pasta_saida / nome_arquivo
        
        # Cria o workbook e a planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Geral"
        
        # Adiciona o título
        ws.merge_cells("A1:K1")
        ws["A1"] = "Relação de Instrumentos"
        ws["A1"].font = self.fonte_titulo
        ws["A1"].alignment = self.alinhamento_central
        
        # Adiciona os cabeçalhos
        colunas = [
            "SPG",
            "Ensaio",
            "Instrumento",
            "Tag",
            "Localização",
            "Faixa",
            "Unidade",
            "Classe",
            "Última Calibração",
            "Próxima Calibração",
            "Status"
        ]
        
        for col, nome in enumerate(colunas, 1):
            celula = ws.cell(row=2, column=col)
            celula.value = nome
            celula.font = self.fonte_cabecalho
            celula.fill = self.preenchimento_cabecalho
            celula.alignment = self.alinhamento_central
            celula.border = self.borda_fina
        
        # Adiciona os dados
        for row, (_, linha) in enumerate(dados.iterrows(), 3):
            for col, valor in enumerate(linha, 1):
                celula = ws.cell(row=row, column=col)
                celula.value = valor
                celula.font = self.fonte_dados
                celula.alignment = self.alinhamento_central
                celula.border = self.borda_fina
        
        # Ajusta as larguras das colunas
        for col in range(1, len(colunas) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Adiciona filtros
        ws.auto_filter.ref = f"A2:K{len(dados) + 2}"
        
        # Congela o cabeçalho
        ws.freeze_panes = "A3"
        
        # Salva o arquivo
        wb.save(caminho_arquivo)
        
        return caminho_arquivo
    
    def criar_planilha_ensaio(
        self,
        dados: pd.DataFrame,
        spg: str,
        ensaio: str
    ) -> Path:
        """
        Cria a planilha de um ensaio específico.
        
        Args:
            dados: DataFrame com os dados dos instrumentos.
            spg: SPG do ensaio.
            ensaio: Nome do ensaio.
            
        Returns:
            Caminho do arquivo Excel gerado.
        """
        # Cria o arquivo Excel
        nome_arquivo = f"[{spg}] {ensaio} - Relação de Instrumentos.xlsx"
        caminho_arquivo = self.pasta_saida / nome_arquivo
        
        # Cria o workbook e a planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Instrumentos"
        
        # Adiciona o título
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Relação de Instrumentos - {spg} - {ensaio}"
        ws["A1"].font = self.fonte_titulo
        ws["A1"].alignment = self.alinhamento_central
        
        # Adiciona os cabeçalhos
        colunas = [
            "SPG",
            "Ensaio",
            "Instrumento",
            "Tag",
            "Localização",
            "Faixa",
            "Unidade",
            "Classe",
            "Última Calibração",
            "Próxima Calibração",
            "Status"
        ]
        
        for col, nome in enumerate(colunas, 1):
            celula = ws.cell(row=2, column=col)
            celula.value = nome
            celula.font = self.fonte_cabecalho
            celula.fill = self.preenchimento_cabecalho
            celula.alignment = self.alinhamento_central
            celula.border = self.borda_fina
        
        # Adiciona os dados
        for row, (_, linha) in enumerate(dados.iterrows(), 3):
            for col, valor in enumerate(linha, 1):
                celula = ws.cell(row=row, column=col)
                celula.value = valor
                celula.font = self.fonte_dados
                celula.alignment = self.alinhamento_central
                celula.border = self.borda_fina
                
                # Aplica formatação condicional
                if col == 11:  # Coluna Status
                    if valor == "ATIVO":
                        celula.fill = self.preenchimento_verde
                    elif valor == "BLOQUEADO":
                        celula.fill = self.preenchimento_vermelho
                
                elif col in [9, 10]:  # Colunas de calibração
                    if pd.isna(valor):
                        celula.fill = self.preenchimento_vermelho
                    else:
                        data = pd.to_datetime(valor)
                        if data < datetime.now():
                            celula.fill = self.preenchimento_vermelho
                        elif (data - datetime.now()).days <= 30:
                            celula.fill = self.preenchimento_amarelo
        
        # Ajusta as larguras das colunas
        for col in range(1, len(colunas) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Adiciona filtros
        ws.auto_filter.ref = f"A2:K{len(dados) + 2}"
        
        # Congela o cabeçalho
        ws.freeze_panes = "A3"
        
        # Salva o arquivo
        wb.save(caminho_arquivo)
        
        return caminho_arquivo 